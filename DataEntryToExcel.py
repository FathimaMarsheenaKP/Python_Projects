from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import Label, Entry, Button
################################

class DataEntryForm(tk.Tk):
    def __init__(self):
        super().__init__()


        self.title("Data Entry Form")
        self.geometry("500x200")

        self.item_label = Label(self, text = 'Item')
        self.item_entry = Entry(self)

        self.quantity_label = Label(self, text = 'Quantity')
        self.quantity_entry = Entry(self)

        self.price_label = Label(self, text = 'Price')
        self.price_entry = Entry(self)

        self.add_btn = Button(self, text = 'Add to Excel', command = self.add_data)

        self.item_label.grid(row = 0, column = 0, padx=10, pady=10)
        self.item_entry.grid(row = 0, column = 1, padx=10, pady=10)

        self.quantity_label.grid(row = 1, column = 0, padx=10, pady=10)
        self.quantity_entry.grid(row = 1, column = 1, padx=10, pady=10)

        self.price_label.grid(row = 2, column = 0, padx=10, pady=10)
        self.price_entry.grid(row = 2, column = 1, padx=10, pady=10)

        self.add_btn.grid(row = 3, column = 2, padx=10, pady=10)

    def add_data(self):
        item = self.item_entry.get()
        quantity = self.quantity_entry.get()
        price = self.price_entry.get()

        if item and quantity and price:
            self.write_to_excel(item, quantity, price)

        self.item_entry.delete(0,'end')
        self.quantity_entry.delete(0,'end')
        self.price_entry.delete(0,'end')

    def write_to_excel(self, item, quantity, price):
        try:
            workbook = load_workbook('item.xlsx')
        except FileNotFoundError:
            workbook = Workbook()

        sheet = workbook.active

        row = [item, quantity, price]
        sheet.append(row)

        workbook.save('item.xlsx')





app = DataEntryForm()
app.mainloop()



